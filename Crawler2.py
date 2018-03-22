import requests
import os
import time
import datetime
import multiprocessing
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl import load_workbook

# *************************************************
# Crawler2 : 只有一个excel文件，所有城市每个月的数据一张sheet
# *************************************************

JJJCITIES = ['北京市', '天津市', '石家庄市', '唐山市',
             '秦皇岛市', '邯郸市', '邢台市', '保定市',
             '张家口市', '承德市', '沧州市', '廊坊市', '衡水市']


# get the web page content
def get_page(page_number, date):
    request_headers = {
        'Host': 'datacenter.mep.gov.cn:8099',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:57.0) Gecko/20100101 Firefox/57.0',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
        'Accept-Language': 'en-US,en;q=0.5',
        'Accept-Encoding': 'gzip, deflate',
        'Referer': 'http://datacenter.mep.gov.cn:8099/ths-report/report!list.action',
        'Content-Type': 'application/x-www-form-urlencoded',
        'Content-Length': '10370',
        'Cookie': 'JSESSIONID=2F3F30CB410BF78BEF91B308C2248DCE; _gscu_95520784=11958468yzg6r511; wdcid=1a3e2c5b5693eede; Hm_lvt_2be83df269939e306c52614d37829e53=1514027024,1514076684,1514098113,1514115383; JSESSIONID=6BB09BD93B9E0A0C6A5BA79F6FDD1533',
        'Connection': 'keep-alive',
        'Upgrade-Insecure-Requests': '1',
        'Cache-Control': 'max-age=0'
    }

    # request_url = 'http://datacenter.mep.gov.cn:8099/ths-report/report!list.action?xmlname=1462259560614'
    request_url = 'http://datacenter.mep.gov.cn:8099/ths-report/report!list.action?xmlname=1462259560614'

    params = {
        'page.pageNo': str(page_number),
        'page.orderBy': '',
        'page.order': '',
        'gisDataJson': [
            {"GRADE": "一级", "ROWNUM": "31", "ID": "61FAEAFF389B673DE050007F01001E82", "AQI": "28", "CITY": "呼伦贝尔市",
             "STATUS": "优", "MAIN_POLLUTANT": "", "CITYCODE": "150700", "OPER_DATE": "2018-01-04"},
            {"GRADE": "二级", "ROWNUM": "32", "ID": "61FAEAFF389C673DE050007F01001E82", "AQI": "54", "CITY": "巴彦淖尔市",
             "STATUS": "良", "MAIN_POLLUTANT": "PM10", "CITYCODE": "150800", "OPER_DATE": "2018-01-04"},
            {"GRADE": "一级", "ROWNUM": "33", "ID": "61FAEAFF389D673DE050007F01001E82", "AQI": "45", "CITY": "乌兰察布市",
             "STATUS": "优", "MAIN_POLLUTANT": "", "CITYCODE": "150900", "OPER_DATE": "2…级", "ROWNUM": "58",
             "ID": "61FAEAFF38B6673DE050007F01001E82", "AQI": "51", "CITY": "松原市", "STATUS": "良",
             "MAIN_POLLUTANT": "PM10",
             "CITYCODE": "220700", "OPER_DATE": "2018-01-04"},
            {"GRADE": "一级", "ROWNUM": "59", "ID": "61FAEAFF38B7673DE050007F01001E82", "AQI": "50", "CITY": "白城市",
             "STATUS": "优", "MAIN_POLLUTANT": "", "CITYCODE": "220800", "OPER_DATE": "2018-01-04"},
            {"GRADE": "二级", "ROWNUM": "60", "ID": "61FAEAFF38B8673DE050007F01001E82", "AQI": "52", "CITY": "延边州",
             "STATUS": "良", "MAIN_POLLUTANT": "PM2.5", "CITYCODE": "222400", "OPER_DATE": "2018-01-04"}],
        'isdesignpatterns': 'false',
        'CITY': '',
        # 'V_DATE': '2018-01-04',
        # 'E_DATE': '2018-01-04'
        'V_DATE': str(date),
        'E_DATE': str(date)
    }

    r = requests.post(request_url, params, headers=request_headers)
    print(r.status_code)
    # print(r.raise_for_status())
    # print(r.content)
    return r.content


[['1', '8B23B8B1B5364378802AF2BEDE2DE3C5', '北京市', '248', 'PM2.5', '五级', '2017-01-02', '110000', '重度污染'],
 ['2', 'C1E9A1F87F604792AEE73EB8103503F6', '天津市', '330', 'PM2.5', '六级', '2017-01-02', '120000', '严重污染'],
 ['3', 'C849C19624E54A7D885189F1EF47B272', '石家庄市', '392', 'PM10', '六级', '2017-01-02', '130100', '严重污染'],
 ['4', '0B172A15803340639A7181F713A12F85', '唐山市', '303', 'PM2.5', '六级', '2017-01-02', '130200', '严重污染'],
 ['5', '11944CD37EA443088E0EBA91E9419E7A', '秦皇岛市', '166', 'PM2.5', '四级', '2017-01-02', '130300', '中度污染'],
 ['6', '4A16E1C08DBC4C7AB1F1B00C0C5E148D', '邯郸市', '360', 'PM10', '六级', '2017-01-02', '130400', '严重污染'],
 ['7', 'A0AEE64697EE4EBCA37BA2BC22ABB9E3', '邢台市', '355', 'PM10', '六级', '2017-01-02', '130500', '严重污染'],
 ['8', '0589A9C2BE114DD68205B13C85D2E16C', '保定市', '288', 'PM2.5', '五级', '2017-01-02', '130600', '重度污染'],
 ['9', '0CE83558F9D84826BBAE75C2EF45DA3F', '承德市', '93', 'PM10', '二级', '2017-01-02', '130800', '良'],
 ['10', 'A14DC5A63C234F608BB1EA17B388CA9B', '沧州市', '229', 'PM2.5', '五级', '2017-01-02', '130900', '重度污染'],
 ['11', 'A7D03791CB274FF6A2754B0F7F1CA809', '廊坊市', '328', 'PM2.5', '六级', '2017-01-02', '131000', '严重污染'],
 ['12', 'A024B7C913004100BFE70165C3810920', '衡水市', '252', 'PM10', '五级', '2017-01-02', '131100', '重度污染'],
 ['13', '280C4BB64DCC4213806D17297A7E21A6', '张家口市', '62', 'PM10', '二级', '2017-01-02', '131200', '良']
 ]


# parse jjj's page content into a data list
def parse_jjj_data(content):
    soup = BeautifulSoup(content, "html.parser")
    # print(soup.prettify())

    all_data = []
    for row_number in range(30):
        row = []
        for col_number in range(9):
            cell = soup.find('td', rowid=str(row_number + 1), colid=str(col_number))
            if cell:
                row.append(cell.text)
                # print(cell.text)
        all_data.append(row)
        # print('-------------------')

    jjj_data = []
    for row in all_data:
        global JJJCITIES
        if row[2] in JJJCITIES:
            jjj_data.append(row)
    return jjj_data


def get_calculated_data():
    bj, tj, sjz, ts, qhd, hd, xt, bd, zjk, cd, cz, lf, hs = []
    cityname_to_list = {'北京市': bj, '天津市': tj, '石家庄市': sjz, '唐山市': ts,
                        '秦皇岛市': qhd, '邯郸市': hd, '邢台市': xt, '保定市': bd,
                        '张家口市': zjk, '承德市': cd, '沧州市': cz, '廊坊市': lf, '衡水市': hs}

    oneday = datetime.timedelta(days=1)
    start_date = datetime.date(2017, 1, 1)

    date = start_date

    while int(date.year) != 2018:
        content = get_page(1, date.__str__())
        jjj_data = parse_jjj_data(content)
        print(jjj_data)

        for row in jjj_data:
            one_day_data = [row[2], row[3], row[6]]
            cityname_to_list[row[2]].append(one_day_data)

        date += oneday



# write th data into xlsx file
def write_data(data):
    col_name = ['Serial_Num', 'Long_Code', 'City_Name', 'AQI', 'Main_Pollutant', 'City_Class', 'Date', 'Division_Code',
                'Pollution_Class']

    dest_filename = data[0][6][0:4]
    # dest_sheetname = data[data_row_num][6][0:7]


    # if there isn't a dest_file, then create one
    if not os.path.exists(dest_filename + '.xlsx'):
        wb = Workbook()
        wb.save(dest_filename + '.xlsx')
        print('create file: ' + wb.path)

    wb = load_workbook(dest_filename + '.xlsx')


def get_one_year_data(start_date):
    oneday = datetime.timedelta(days=1)

    date = start_date
    end_year = int(start_date.year) + 1

    prcname = multiprocessing.current_process().name
    while int(date.year) != end_year:
        print('---进程:' + prcname + '--- get data at ' + date.__str__())

        content = get_page(1, date.__str__())
        data = parse_jjj_data(content)
        print(data)

        write_data(data)

        date += oneday
