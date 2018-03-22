import requests
import os
import time
import datetime
import multiprocessing
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl import load_workbook

# *************************************************
# Crawler1 : 每个城市一个excel文件，每个月的数据一张sheet
# *************************************************

# get the web page content
def get_page(page_number, date):
    request_headers = {
        'Host': 'datacenter.mep.gov.cn:8099',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:57.0) Gecko/20100101 Firefox/57.0',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
        'Accept-Language': 'en-US,en;q=0.5',
        'Accept-Encoding': 'gzip, deflate',
        'Referer': 'http://datacenter.mep.gov.cn:8099/ths-report/report!list.action',
        'Content-Type': 'application/x-www-form-urlencoded',
        'Content-Length': '10370',
        'Cookie': 'JSESSIONID=2F3F30CB410BF78BEF91B308C2248DCE; _gscu_95520784=11958468yzg6r511; wdcid=1a3e2c5b5693eede; Hm_lvt_2be83df269939e306c52614d37829e53=1514027024,1514076684,1514098113,1514115383; JSESSIONID=6BB09BD93B9E0A0C6A5BA79F6FDD1533',
        'Connection': 'keep-alive',
        'Upgrade-Insecure-Requests': '1',
        'Cache-Control': 'max-age=0'
    }

    # request_url = 'http://datacenter.mep.gov.cn:8099/ths-report/report!list.action?xmlname=1462259560614'
    request_url = 'http://datacenter.mep.gov.cn:8099/ths-report/report!list.action?xmlname=1462259560614'

    params = {
        'page.pageNo': str(page_number),
        'page.orderBy': '',
        'page.order': '',
        'gisDataJson': [
            {"GRADE": "一级", "ROWNUM": "31", "ID": "61FAEAFF389B673DE050007F01001E82", "AQI": "28", "CITY": "呼伦贝尔市",
             "STATUS": "优", "MAIN_POLLUTANT": "", "CITYCODE": "150700", "OPER_DATE": "2018-01-04"},
            {"GRADE": "二级", "ROWNUM": "32", "ID": "61FAEAFF389C673DE050007F01001E82", "AQI": "54", "CITY": "巴彦淖尔市",
             "STATUS": "良", "MAIN_POLLUTANT": "PM10", "CITYCODE": "150800", "OPER_DATE": "2018-01-04"},
            {"GRADE": "一级", "ROWNUM": "33", "ID": "61FAEAFF389D673DE050007F01001E82", "AQI": "45", "CITY": "乌兰察布市",
             "STATUS": "优", "MAIN_POLLUTANT": "", "CITYCODE": "150900", "OPER_DATE": "2…级", "ROWNUM": "58",
             "ID": "61FAEAFF38B6673DE050007F01001E82", "AQI": "51", "CITY": "松原市", "STATUS": "良",
             "MAIN_POLLUTANT": "PM10",
             "CITYCODE": "220700", "OPER_DATE": "2018-01-04"},
            {"GRADE": "一级", "ROWNUM": "59", "ID": "61FAEAFF38B7673DE050007F01001E82", "AQI": "50", "CITY": "白城市",
             "STATUS": "优", "MAIN_POLLUTANT": "", "CITYCODE": "220800", "OPER_DATE": "2018-01-04"},
            {"GRADE": "二级", "ROWNUM": "60", "ID": "61FAEAFF38B8673DE050007F01001E82", "AQI": "52", "CITY": "延边州",
             "STATUS": "良", "MAIN_POLLUTANT": "PM2.5", "CITYCODE": "222400", "OPER_DATE": "2018-01-04"}],
        'isdesignpatterns': 'false',
        'CITY': '',
        # 'V_DATE': '2018-01-04',
        # 'E_DATE': '2018-01-04'
        'V_DATE': str(date),
        'E_DATE': str(date)
    }

    r = requests.post(request_url, params, headers=request_headers)
    print(r.status_code)
    # print(r.raise_for_status())
    # print(r.content)
    return r.content


# parse the page content into a data list
def parse_data(content):
    soup = BeautifulSoup(content, "html.parser")
    # print(soup.prettify())

    data = []
    for row_number in range(30):
        row = []
        for col_number in range(9):
            cell = soup.find('td', rowid=str(row_number + 1), colid=str(col_number))
            if cell:
                row.append(cell.text)
                # print(cell.text)
        data.append(row)
        # print('-------------------')

    return data


# write th data into xlsx file
def write_data(data):
    col_name = ['Serial_Num', 'Long_Code', 'City_Name', 'AQI', 'Main_Pollutant', 'City_Class', 'Date', 'Division_Code',
                'Pollution_Class']

    for data_row_num in range(0, 30):
        if data[data_row_num]:
            dest_filename = 'InitData/' + data[data_row_num][2]
            dest_sheetname = data[data_row_num][6][0:7]

            if '市' not in dest_filename:
                print('\'市\' is not in dest_filename : ' + dest_filename)
                break

            # if there isn't a dest_file, then create one
            if not os.path.exists(dest_filename + '.xlsx'):
                wb = Workbook()
                wb.save(dest_filename + '.xlsx')
                print('create file: ' + wb.path)

            wb = load_workbook(dest_filename + '.xlsx')

            # if there isn't a dest_sheet, then create one
            if dest_sheetname not in wb:
                ws = wb.create_sheet(dest_sheetname)
                print('create sheet: ' + ws.title)

            ws = wb.get_sheet_by_name(dest_sheetname)

            if ws.max_row == 1:
                for col_number in range(1, 10):
                    ws.cell(row=1, column=col_number).value = col_name[col_number - 1]

            row_number = ws.max_row + 1
            for col_number in range(1, 10):
                ws.cell(row=row_number, column=col_number).value = data[data_row_num][col_number - 1]

            wb.save(dest_filename + '.xlsx')


def get_one_year_data(start_date):
    oneday = datetime.timedelta(days=1)

    date = start_date
    end_year = int(start_date.year) + 1

    # prcname = multiprocessing.current_process().name
    while int(date.year) != end_year:
        # print('---进程:' + prcname + '--- get data at ' + date.__str__())

        content = get_page(1, date.__str__())
        data = parse_data(content)
        print(data)

        write_data(data)

        date += oneday


if __name__ == "__main__":
    start_date1 = datetime.date(2014, 1, 1)
    start_date2 = datetime.date(2015, 1, 1)
    start_date3 = datetime.date(2016, 1, 1)
    start_date4 = datetime.date(2017, 1, 1)

    get_one_year_data(start_date4)
