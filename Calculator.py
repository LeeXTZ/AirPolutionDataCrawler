from openpyxl import Workbook
from openpyxl import load_workbook
import os

JJJCITIES = ['北京市', '天津市', '石家庄市', '唐山市',
             '秦皇岛市', '邯郸市', '邢台市', '保定市',
             '张家口市', '承德市', '沧州市', '廊坊市', '衡水市']
dest_filename = 'AirPollutionData'

if not os.path.exists(dest_filename + '.xlsx'):
    wb = Workbook()
    wb.save(dest_filename + '.xlsx')
    print('create file: ' + wb.path)

dest_wb = load_workbook(dest_filename + '.xlsx')
dest_ws = dest_wb.active

title = ['Admin_Name', '2017-01', '2017-02', '2017-03', '2017-04', '2017-05', '2017-06', '2017-07', '2017-08',
         '2017-09',
         '2017-10', '2017-11', '2017-12']

for c in range(0, len(title)):
    dest_ws.cell(row=1, column=c + 1).value = title[c]

data = []

# data.append(title)

for city in JJJCITIES:
    wb = load_workbook('InitData/' + city + '.xlsx')
    mean_list = [city]

    for month in range(1, 13):
        if month < 10:
            month_str = '0' + str(month)
        else:
            month_str = str(month)
        month_str = '2017-' + month_str

        if month_str not in wb:
            print(city + ': missing sheet of' + month_str)
            break
        else:
            ws = wb.get_sheet_by_name(month_str)
            api_column = ws['D']

            sum = 0
            mean = 0
            if len(api_column) > 1:
                for i in range(1, len(api_column)):
                    sum += int(api_column[i].value)
                mean = sum / (len(api_column) - 1)
                mean_list.append(mean)
            else:
                print(city + ': missing data of' + month_str)

    data.append(mean_list)

for row in data:
    dest_ws.append(row)

dest_wb.save(dest_filename + '.xlsx')
